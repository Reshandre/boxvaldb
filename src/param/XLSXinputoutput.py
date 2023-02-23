from openpyxl import Workbook
from os import (listdir,path)
from pathlib import Path
from param import settings
from shutil import copyfile

from tools.tools_global import getModel

from .settings import (
    DATA_LOAD_DUMP_BACKUP,
    DATA_LOAD_DUMP_DIR,
    DATA_LOAD_DUMP_EXPORT_SUFFIX,
    DATA_LOAD_DUMP_EXTENTION, 
    DATA_LOAD_DUMP_IMPORT_SUFFIX, )

from .tools import ( 
    getLoadDataImportPath,
    uniqueFieldConstraintslist,
    getLoadDataExportPath,
    getRawColumns, listForeignKeys,
    loadData,
    mapColumns,
    mapRawColumns, 
    composeColumns,
    openWorkBook,
    readColumns,
    buildWorkBook,
)


def saveWorkBook(wb,fileName):
    wb.save(fileName)
    
def deleteOthereSheets(wb,sheetName):
    for existingSheetName in wb.sheetnames:
        if existingSheetName != sheetName:
            wb.remove(wb[existingSheetName])

def runExport(appName, modelName):
    pass
    fileName = getLoadDataExportPath(appName, modelName)
    print(f"Export FilePath {fileName} /n appName {appName} modelName {modelName}")
    wb = Workbook()
    wb.create_sheet(modelName)
    deleteOthereSheets(wb,modelName)
    modelObject = getModel(appName, modelName)
    rawColumns = getRawColumns(modelObject)
    columns = mapRawColumns(rawColumns,appName, modelName)
    composeColumns(wb,columns,modelName)
    buildWorkBook(wb, columns,rawColumns, appName, modelName)
    saveWorkBook(wb, fileName)

    

    

def runImport(appName, modelName):
    """Batch import of excel named after the names of the app and model

    Args:
        appName (string): Name of the app
        modelName (string): Name of the model
    """
    fileName = getLoadDataImportPath(appName, modelName)
    print(f"Import FileName {fileName} /n appName {appName} modelName {modelName}")
    wb = openWorkBook(fileName)
    columns = readColumns(wb, modelName)
    mappedColumns = mapColumns(columns, appName, modelName)
    loadData(wb, mappedColumns, appName, modelName)
    return



def runBackup(appName, modelName):
    archiveDir = DATA_LOAD_DUMP_DIR / appName / DATA_LOAD_DUMP_BACKUP
    importExportDir = DATA_LOAD_DUMP_DIR / appName 
    importFilePath = getLoadDataImportPath(appName, modelName)
    exportFilePath = getLoadDataExportPath(appName, modelName)
    lst = listdir(archiveDir)
    numList= [f.split('_')[1].split('.')[0] for f in lst if '_' in f and modelName in f and f.split('_')[0] ==modelName]
    maximum = 0
    if len(numList) > 0:
        maximum = max(numList)
    fileNumber = int(maximum) + 1
    baseFileNameToBeSavedTarget = f"{modelName}_{fileNumber}{DATA_LOAD_DUMP_EXTENTION}" 
    PathFileToBeSavedTarget = archiveDir / baseFileNameToBeSavedTarget
    if path.isfile(importFilePath):
        copyfile(importFilePath,PathFileToBeSavedTarget)
    copyfile(exportFilePath,importFilePath)
    
    
    
    
    





