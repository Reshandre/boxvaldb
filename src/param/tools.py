
from django.apps import apps

from tools.tools_global import getForeignKeyValue, getModel
from .models import ModelFieldMapper
from tools.tools_global import listForeignKeys


from django.conf import settings as gsettings
from .settings import (
    DATA_LOAD_DUMP_DIR,
    DATA_LOAD_DUMP_EXPORT_SUFFIX,
    DATA_LOAD_DUMP_EXTENTION, 
    DATA_LOAD_DUMP_IMPORT_SUFFIX, )
from importlib import import_module

from param.models import ModelFieldMapper
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import Fill
from openpyxl.styles.fills import DEFAULT_GRAY_FILL,FILL_SOLID

from openpyxl import load_workbook, Workbook

"""Functions in support of the param app

 
 """



def mapColumns(columns, appName, modelName):
    """Columns of the excel sheet are transformed into the names of the model

    Args:
        columns (list of strings): Columns of the excel sheet
        appName (string): app name
        modelName (string): model name 

    Returns:
        list of string: list of model characteristics
    """
    mappedColumns = []
    for c in columns:
        modelFieldName = c
        qs = ModelFieldMapper.objects.filter(
            AppName=appName, ModelName=modelName, ModelMappedName=c)
        qsFound = len(qs) == 1
        if qsFound:
            modelFieldName = qs[0].ModelFieldName
        mappedColumns.append(modelFieldName)
    return mappedColumns


def composeColumns(wb, columns, modelName):
    """Sets the header in the first row

    Args:
        wb (WorkBook): WorkBook excel
        columns (_type_): _description_
        modelName (_type_): _description_
    """
    ws = wb[modelName]
    ft=Font(bold=True)
    ws.append(columns)
    # ft = Font(bold=True)
    for cell in ws[1:1]:
        cell.font = ft


def mapRawColumns(rawColumns, appName, modelName):
    """Replace name of importable fields of the model to the heading of the excel sheet

    Args:
        rawColumns (list of string): list of columns of the model
        appName (string): app name
        modelName (string): model name

    Returns:
        list of string: list of fields who become the list of the excel sheet first row
    """
    mappedColumns = []
    noExportList = getNoExportList(appName,modelName)
    for cr in rawColumns:
        modelMappedName = cr # in case not a mapped field
        if cr in noExportList:
            continue
        qs = ModelFieldMapper.objects.filter(
            AppName=appName, ModelName=modelName, ModelFieldName=cr)
        qsFound = len(qs) == 1
        if qsFound and len(qs[0].ModelMappedName) >0:
            modelMappedName = qs[0].ModelMappedName
        mappedColumns.append(modelMappedName)
    return mappedColumns


# 




def openWorkBook(fileName):
    wb = load_workbook(filename=fileName)
    return wb





def loadData(wb, mappedColumns, appName, modelName):
    modelObject = getModel(appName, modelName)
    lstForeignKeys = listForeignKeys(modelObject, appName, modelName)
    sheet = wb[modelName]
    constraints = uniqueFieldConstraintslist(modelObject, appName, modelName)
    for count, r in enumerate(sheet.rows):
        if count == 0:
            continue
        rowData = {}
        for indx, c in enumerate(mappedColumns):
            if r[indx].value == None:
                rowData[c] = None
            else:
                rowData[c] = r[indx].value
                if c in lstForeignKeys:
                    rowData[c] = getForeignKeyValue(
                        appName, modelName, c, rowData[c])
        # constraintsCache = constraints
        # aConstraintList = constraints[0]
        constraintKeys = {}
        constraintCollected=False
        for cns in constraints:
            if constraintCollected:
                break
            for cnsi in cns:
                if constraintCollected:
                    break
                constraintKeys = {}
                for cnsfld in cnsi:                    
                    if cnsfld in rowData.keys():
                        constraintKeys[cnsfld] = rowData[cnsfld]
                        if len(constraintKeys) == len(cnsi):
                            constraintCollected=True
                            break
                    else:
                        break
        # kwargs = {k: v for k, v in rowData.items() if k in aConstraintList}
        deltaDicUpdate = {}
        deltaDicUpdate['defaults'] = {
            k: v for k, v in rowData.items() if k not in constraintKeys}
        deltaDicCreate={k: v for k, v in rowData.items() if k not in constraintKeys}
        kwargsUpdate = constraintKeys | deltaDicUpdate
        kwargsCreate = constraintKeys | deltaDicCreate
        
        # module = import_module(f"{appName}.models")
        qs = modelObject.objects.filter(**constraintKeys)
        if (len(qs) > 0):
            kwargsCreate["id"] = qs[0].id
            o = modelObject(**kwargsCreate)
            o.save()
            pass
        else:
            modelObject.objects.create(**kwargsCreate)
            pass
        
        
        # obj, created = modelObject.objects.update_or_create(**kwargs)
        pass

def getNoExportList(appName,modelName):
        qs_noExportList = ModelFieldMapper.objects.filter(AppName=appName,ModelName=modelName,ModelDoNotExport = True)
        noExportList = [n.ModelFieldName for n in qs_noExportList]
        return noExportList

def fetchForeignKeyCode (id,fieldName,appName, modelName):
    qs = ModelFieldMapper.objects.filter(AppName=appName,ModelName=modelName,ModelFieldName = fieldName)
    if len(qs) == 0:
        resultValue = id
    else:
        fkAppForeignKeyName = qs[0].AppForeignKeyName
        fkModelForeignKeyModel = qs[0].ModelForeignKeyModel
        fkModelForeignKeyField =qs[0].ModelForeignKeyField
        fkModelObject = getModel(fkAppForeignKeyName, fkModelForeignKeyModel)
        if id is not None:
            fkQs = fkModelObject.objects.get(id=id)
            resultValue = eval(f"fkQs.{fkModelForeignKeyField}")
        else:
            resultValue = ''
    return resultValue
    
    
     
    
    
def buildWorkBook(wb, columns,rawColumns, appName, modelName):
    modelObject = getModel(appName, modelName)
    lstForeignKeys = listForeignKeys(modelObject,appName,modelName)
    ws = wb[modelName]
    constraints = uniqueFieldConstraintslist(modelObject, appName, modelName)
    qs = modelObject.objects.all()
    noExportList = getNoExportList(appName,modelName)
    for rownum,r in enumerate(qs):
        row = []
        for f in r.__dict__:
            fieldName = f
            if 'id' != f and'id' in f and f.split('_')[1]=='id':
                fieldName = f.split('_')[0]       
            if fieldName in  noExportList:
                continue
            if fieldName not in rawColumns:
                continue
            if fieldName not in  lstForeignKeys:
                row.append(r.__dict__[f])
            else:
                row.append(fetchForeignKeyCode(r.__dict__[f],fieldName,appName, modelName))
        ws.cell(row=rownum+1,column=1)
        ws.append(row)
        
       
                

            

def readColumns(wb, modelName):
    sheet = wb[modelName]
    columns = sheet[1]
    return [c.value for c in columns]


# def composeColumns(wb, columns):
#     pass


def getLoadData(appName):
    return gsettings.DATA_LOAD_DUMP_DIR / appName


def getLoadDataImportPath(appName, modelName):
    fileName = modelName + DATA_LOAD_DUMP_IMPORT_SUFFIX + \
        DATA_LOAD_DUMP_EXTENTION
    return DATA_LOAD_DUMP_DIR / appName / fileName


def getLoadDataExportPath(appName, modelName):
    fileName = modelName + DATA_LOAD_DUMP_EXPORT_SUFFIX + \
        DATA_LOAD_DUMP_EXTENTION
    return DATA_LOAD_DUMP_DIR / appName / fileName


def uniqueFieldConstraintslist(modelObject, appName, modelName):
    """Get the list of unique field definitions from the model

    Args:
        modelObject (class of model): Class of model
        appName (string): app name
        modelName (string): string name of the mmodel

    Returns:
        list of strings in list: list of fields representing constrains
    """    
    """ """
    uniqueFieldConstraintlst = []
    for fld in modelObject._meta.local_concrete_fields:
        if fld.unique:
            if '_' in fld.name:
                if fld.name.split('_')[1] == 'ptr':
                    continue
            cnstr = [fld.name]
            uniqueFieldConstraintlst.append(cnstr)
    
    for ol in modelObject.__mro__:
        if '__module__' not in ol.__dict__.keys():
            continue
        if ol.__dict__['__module__'].split('.')[0]==appName:
            for oc in ol._meta.constraints:
                cnstr = [oc.fields]
                uniqueFieldConstraintlst.append(cnstr)

    return uniqueFieldConstraintlst

def translateDict(aDict,appName, modelName):
    qs = ModelFieldMapper.objects.filter(
            AppName=appName, ModelName=modelName).values_list('AppName','ModelName','ModelFieldName','ModelMappedName')
    mapper = {e[2]:e[3] for e in qs if e[3]!= ""}
    mapperKeys = mapper.keys()
    returnDict = {}
    for k,d in aDict.items():
        if k in mapperKeys:
            newKey = mapper[k]
        else:
            newKey = k
        returnDict[newKey] = d
        
    return returnDict


def buildModelOutput(appName,modelName,fieldList,filter):
    """Returns a dictionary with a record of the model filtered with 'filter'

    Args:
        appName (string): project app name
        modelName (string): model name of the app
        fieldList (list): selection of fields to be extracted from teh record
        filter (Q object): Filters the record from the model

    Returns:
        Dictionary: LIst of fields
    """
    theModel = getModel(appName,modelName)
    qsEntries = theModel.objects.filter(filter)
    entriesDictTemp = qsEntries.values()
    entriesDictRenamedCols=[]
    for o in entriesDictTemp:    
        entriesDict = {k:o[k] for k in o.keys() if k in fieldList}
        entriesDictRenamedCols.append(translateDict(entriesDict,appName, modelName))    
    return entriesDictRenamedCols
