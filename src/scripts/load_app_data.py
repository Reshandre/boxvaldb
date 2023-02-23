from os import path
from boxvaldb.settings import BASE_DIR
from csv import reader
from openpyxl import load_workbook
import logging
import datetime
import pytz
from tools import script_tools

class runCount:
    def __init__(self):
        self.modelName = ''
        self.updates = 0
        self.inserts = 0
    def startCounting(self,model):
        self.modelName = model.__name__
        self.updates = 0
        self.inserts = 0    
    def updateIncrement(self):
        self.updates += 1
    def insertIncrement(self):
        self.inserts += 1
    def getUpdates(self):
        return self.updates
    def getInserts(self):
        return self.inserts

loadCount = runCount()





def saveWithPk(model,params):
    
    try:
        unique_fields = script_tools.alternateKeys(model)
        search_entries = {k: params[k] for k in unique_fields}
    except:
        search_entries = {}
    if len(search_entries) == 0:
           for f in model._meta.fields:
                if f.unique:
                    if f.name in params.keys():
                        if params[f.name] is not None and  len(params[f.name]) > 0:
                            search_entries[f.name] = params[f.name]
                            break
    if len(search_entries) == 0:
        search_entries = params

    record = model.objects.filter(**search_entries)
    if len(record) >1:
        mes = f' {model.__name__} inside saveWithPk {len(record)} records and search entries {search_entries} '
        logging.warning(mes)
    if len(record) == 1:
        pkDict = {'pk':list(record)[0].id}
        params.update(pkDict)
    toSave = model(**params)
    return toSave

def getForeignKeysId(model,params1):
    fks = script_tools.foreignKeys(model)
    params = params1.copy()
    if (len(fks) > 0):
        for fk in fks:
            fkmodel=fk[0]  
            if len(fk[2]) > 0:
                if fk[1] not in params.keys() and all(item in params.keys() for item in fk[2]):
                    minipars = {}
                    for fld in fk[2]:
                        minipars[fld] = params[fld]
                    local_id= fkmodel.objects.filter(**minipars)[0].id if len(fkmodel.objects.filter(**minipars)) > 0 else -1
                    if local_id > -1:
                        params.update({fk[1]:local_id})
                    for fld in fk[2]:
                        params.pop(fld)



    return params

def saveThis(model,params1):                       
    params = getForeignKeysId(model,params1)
    toSave = model(**params)
    saveMode = {}
    try:
        toSave.save()
    except Exception as e:
        if e.__class__.__name__ == 'IntegrityError':
            toSave = saveWithPk(model,params)   
            saveMode = {'force_update':True}
        try:
            toSave.save(**saveMode)
        except Exception as e:
            if e.__class__.__name__ == 'IntegrityError':
                mes = f"{model.__name__} on {datetime.datetime.now(pytz.timezone('Europe/Paris')).strftime('%y-%m-%d %H_%M_%S')} error {e} after update with data {params} and input {params1} "
                logging.error(mes)
        else:
            loadCount.updateIncrement()
            mes = f"{model.__name__} on {datetime.datetime.now(pytz.timezone('Europe/Paris')).strftime('%y-%m-%d %H_%M_%S')} update of   {params} counting {loadCount.getUpdates()} "
            logging.info(mes)
    else:
        loadCount.insertIncrement()
        mes = f"{model.__name__} on {datetime.datetime.now(pytz.timezone('Europe/Paris')).strftime('%y-%m-%d %H_%M_%S')} insert of  {params} with id {model.objects.latest('pk')}, counting {loadCount.getInserts()}"
        logging.info(mes)


def writeCSVmodel(model,fullPath,importType):
    match importType:
        case 'csv':
            csvfile = open(fullPath)
            line=csvfile.readline()
            delim = script_tools.detectDelimitter(line)
            words= line.split(delim) 
            fields=script_tools.cleanupWords(words)
            csviterator = reader(csvfile,delimiter=delim)
            enumerator = enumerate(csviterator)
        case 'xlsx':
            wb = load_workbook(filename=fullPath,read_only=True)
            ws = wb.active
            row = ws[1]
            words = []
            for cell in row:
                words.append(cell.value)
            fields=script_tools.cleanupWords(words)
            iterator = ws.iter_rows(min_row=2)
            enumerator = enumerate(iterator)
            pass

    modelFieldsDictionary = script_tools.fetchFields(model)  
    modelFields = script_tools.allkeys(model)
    for i, buffer in enumerator:
        params = {}         
        for j , fld in  enumerate(fields):
            if fld in modelFields:
                match importType:
                    case 'csv':
                        val = buffer[j]
                    case 'xlsx':
                        val = buffer[j].value                       
                s = script_tools.convertValue(val,modelFieldsDictionary[fld].get('type'))
                params[fld] = s
        saveThis(model,params)

def fetchParameters(*args):
    components = args[0].split(",")
    params = {}
    for c in components:
        par=c.split("=")
        d = {}
        d[par[0]]=par[1]
        params.update(d)
    return params

def prepareModelParameters(*args):
    params= fetchParameters(*args)
    if len(params) == 0:
        raise Exception('The argument list \'--script-args\' must have one \',\'', 'separated entries with app=<app_name>,type=<either csv or xlsx','nameExtention=<string appended at the end of the model names>')
    app_name= params['app']
    input_extention = params['type']
    nameExtention = params ['nameExtention']
    importType = input_extention
    PRE_FULL_PATH = BASE_DIR / app_name / 'fixtures/' 
    print(f'Files expected to be on directory \'{PRE_FULL_PATH}\'')
    EXT=nameExtention
    return (app_name, input_extention,EXT,PRE_FULL_PATH,importType)

def startLogging(model,PRE_FULL_PATH,fileBaseName):
    logFileName = fileBaseName +"_"+datetime.datetime.now().strftime("%y-%m-%d__%H-%M-%S") +'.log'
    FULL_PATH_LOG = PRE_FULL_PATH / 'log' / logFileName
    print(f'INFO: {model.__name__} Logging in {FULL_PATH_LOG}')   
    logging.basicConfig(filename=FULL_PATH_LOG, encoding='utf-8',filemode ='w' ,level=logging.DEBUG,force=True)
    loadCount.startCounting(model)

def run(*args):

    (app_name, input_extention,EXT,PRE_FULL_PATH,importType) = prepareModelParameters(*args)
    models = script_tools.depthSortedModels(app_name)
    for m in models:
        fileBaseName =  m.__name__+EXT
        fileInputName = fileBaseName +'.'+input_extention

        FULL_PATH_INPUT = PRE_FULL_PATH / fileInputName
        if path.exists(FULL_PATH_INPUT):
            startLogging(m,PRE_FULL_PATH,fileBaseName)
            writeCSVmodel(m,FULL_PATH_INPUT,importType)
            mes= f'Model {m.__name__} has updated its entries based on  {fileInputName} with {loadCount.getInserts()} inserts and {loadCount.getUpdates()} updates '
            print(mes)
            logging.info(mes)          
        else:
            mes = f'Model {m.__name__} has no file named {fileInputName}'
            print(mes)
            logging.info(mes)


 