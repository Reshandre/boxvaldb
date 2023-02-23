from os import path
from stat import filemode
from boxvaldb.settings import BASE_DIR
from csv import reader
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import logging
import datetime
from datetime import datetime, timezone
import pytz
from scripts.load_app_data import fetchParameters
from tools import script_tools

class runCount:
    def __init__(self):
        self.modelName = ''
        self.updates = 0
        self.inserts = 0
        self.exports = 0
    def startCounting(self,model):
        self.modelName = model.__name__
        self.updates = 0
        self.inserts = 0    
        self.exports = 0
    def updateIncrement(self):
        self.updates += 1
    def insertIncrement(self):
        self.inserts += 1
    def exportIncrement(self):
        self.exports += 1
    def getUpdates(self):
        return self.updates
    def getInserts(self):
        return self.inserts
    def getExports(self):
        return self.exports

loadCount = runCount()



def exportModel(model,fullPath,importExportType):
    if model.objects.count() == 0:
        mes = f"{model.__name__} at {datetime.datetime.now().strftime('%y-%m-%d %H:%M:%S')} has no entries"
        logging.info(mes)
        print(mes)
    else:
        queryset = model.objects.first()
        keyList = list(queryset.__dict__.keys())
        keyList.pop(0)
        modelFields = keyList              
        modelFieldsDictionary = script_tools.fetchFields(model)  
        match importExportType:
            case 'csv':
                delim = delimitter
                csvfile = open(fullPath,filemode='w',encoding='utf-8')
                line=''
                for i, colName in enumerate(modelFields):
                    line += colName + delim
                csvfile.writeline(line)
            case 'xlsx':
                wb = Workbook()
                ws = wb.active
                ws.title = model.__name__
                for i, colName in enumerate(modelFields):
                    ws.cell(row=1,column=i+1,value=colName)

        queryset = model.objects.all()
        match importExportType:
                case 'csv':
                    for i, queryRow in enumerate(queryset):
                        line = ""
                        loadCount.exportIncrement()
                        for j,cell in enumerate(queryRow):
                            line += cell.value + delim
                        csvfile.close()
                case 'xlsx':

                    for i, queryRow in enumerate(queryset):
                        objectList = list(queryRow.__dict__.values())
                        objectList.pop(0)
                        objectListEnumerator = enumerate(objectList)
                        loadCount.exportIncrement()
                        for j,cell in objectListEnumerator:
                            if type(cell).__name__ == 'UUID':
                                tCell=cell.hex
                            else:
                                tCell = cell
                            thisCell = ws.cell(row=i+2,column=j+1,value=tCell)
                            if modelFields[j].endswith('_id'):
                                thisCell.number_format = 'General'
                            else:    
                                modelCellType = modelFieldsDictionary[modelFields[j]]['type']
                                match modelCellType:
                                    case "CharField":
                                        thisCell.number_format = numbers.FORMAT_TEXT
                                    case "DateField":
                                        thisCell.number_format = 'yy'
                                    case "BigAutoField":
                                        thisCell.number_format = 'General'
                                    case 'ForeignKey':
                                        thisCell.number_format = 'General'
                                    case 'TextField':
                                        thisCell.number_format = numbers.FORMAT_TEXT
                                        thisCell.alignment = Alignment(wrap_text=True)
                                        ws.row_dimensions[i+2].height = 40
                                        # ws.column_dimensions[j+1].width = 80
                                    case 'UUIDField':
                                        thisCell.number_format = numbers.FORMAT_TEXT
                                    case '_':
                                        thisCell.number_format = 'General'
                                        
                    wb.save(fullPath)
    
delimitter = ";"


def prepareModelParameters(*args):
    params= fetchParameters(*args)
    if len(params) == 0:
        raise Exception('The argument list \'--script-args\' must have one \',\'', 'separated entries with app=<app_name>,type=<either csv or xlsx','nameExtention=<string appended at the end of the model names>')
    app_name= params['app']
    input_extention = params['type']
    nameExtention = params ['nameExtention']
    importExportType = input_extention
    PRE_FULL_PATH = BASE_DIR / app_name / 'fixtures/' 
    print(f'Files will be exported on directory \'{PRE_FULL_PATH}\'')
    EXT=nameExtention
    return (app_name, input_extention,EXT,PRE_FULL_PATH,importExportType)

def startLogging(PRE_FULL_PATH,exportType):
    fileBaseName = "ImportTraceOn"+exportType
    logFileName = fileBaseName + '_' + datetime.now(pytz.timezone('Europe/Paris')).strftime("%y-%m-%d_%H_%M_%S") +'.log'
    FULL_PATH_LOG = PRE_FULL_PATH / 'log' / logFileName
    print(f'INFO: Logging in {FULL_PATH_LOG}')   
    logging.basicConfig(filename=FULL_PATH_LOG, encoding='utf-8',filemode ='w' ,level=logging.DEBUG,force=True)
 

def run(*args):

    (app_name, input_extention,EXT,PRE_FULL_PATH,exportType) = prepareModelParameters(*args)
    models = script_tools.depthSortedModels(app_name)
    startLogging(PRE_FULL_PATH,exportType)

    for m in models:
        fileBaseName =  m.__name__+EXT
        fileInputOutputName = fileBaseName +'.'+input_extention
        loadCount.startCounting(m)
        FULL_PATH_INPUT = PRE_FULL_PATH / fileInputOutputName
        exportModel(m,FULL_PATH_INPUT,exportType)
        mes= f"Model {m.__name__} has exported its {loadCount.getExports()} entries at {datetime.now().strftime('%y-%m-%d %H:%M:%S')}"
        print(mes)
        logging.info(mes)          



 